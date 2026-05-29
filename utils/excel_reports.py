"""
Excel hisobotlar — tugab qolayotgan mahsulotlar va sotuv/foyda hisoboti
"""
import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from config import uz_now
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Ranglar
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUB_HEADER_FONT = Font(bold=True, color="1F4E78", size=11)
ZERO_STOCK_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # qizg'ish
LOW_STOCK_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")    # sariq
TOTAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=12)
THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
MONEY_FMT = '#,##0" so\'m"'


def _autosize(ws, min_width=10, max_width=42):
    """Ustun kengligini avtomatik moslash"""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value is not None:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _temp_path(prefix: str) -> str:
    ts = uz_now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(tempfile.gettempdir(), f"{prefix}_{ts}.xlsx")


# ============ TUGAB QOLAYOTGAN MAHSULOTLAR ============

def generate_low_stock_report(products: list) -> str:
    """
    Tugab qolayotgan mahsulotlar uchun Excel.
    Toshkentdan zakaz qilish uchun ro'yxat.

    products: get_low_stock_products() dan kelgan ro'yxat
    Qaytaradi: yaratilgan .xlsx fayl yo'li
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tugab qolganlar"

    # Sarlavha
    ws.merge_cells("A1:G1")
    ws["A1"] = "📉 TUGAB QOLAYOTGAN MAHSULOTLAR — Toshkentdan zakaz uchun"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Sana: {uz_now().strftime('%d.%m.%Y %H:%M')}   |   Manzil: Amudaryo tumani, Mangit shahri"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="555555")

    # Jadval sarlavhasi
    headers = ["#", "Kategoriya", "Brend", "Model", "Mahsulot turi", "Mavjud", "Zarur (min)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Mahsulotlarni kategoriya bo'yicha gruhlash
    row = 5
    n = 1
    current_cat = None
    total_zero = 0
    total_low = 0

    for p in products:
        cat = p.get('category_name', '?')

        # Kategoriya o'zgargan bo'lsa, gruh sarlavhasi
        if cat != current_cat:
            current_cat = cat
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(row=row, column=1, value=f"📂 {cat}")
            cell.font = SUB_HEADER_FONT
            cell.fill = SUB_HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border = BORDER
            row += 1

        qty = int(p.get('quantity', 0) or 0)
        min_qty = int(p.get('min_quantity', 0) or 0)
        need_total = max(min_qty * 2 - qty, min_qty)  # zaxira uchun min*2 ga to'ldirish

        row_data = [
            n,
            cat,
            p.get('brand_name') or "—",
            p.get('model_name') or "—",
            p.get('name', ''),
            qty,
            need_total,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 6, 7) else "left",
                vertical="center"
            )

        # Rang: 0 ta bo'lsa qizil, oz bo'lsa sariq
        fill = ZERO_STOCK_FILL if qty == 0 else LOW_STOCK_FILL
        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).fill = fill

        if qty == 0:
            total_zero += 1
        else:
            total_low += 1

        row += 1
        n += 1

    # Yakuniy qator
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1,
            value=f"JAMI: {n - 1} ta mahsulot  |  Tugab qolgan: {total_zero}  |  Tugayotgan: {total_low}").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    _autosize(ws)
    # Maxsus ustun kengliklari
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14

    path = _temp_path("tugab_qolganlar")
    wb.save(path)
    return path


# ============ SOTUV / FOYDA HISOBOTI ============

def generate_sales_report(summary: dict, by_category: list, details: list,
                          date_from: str, date_to: str) -> str:
    """
    Sotuv va foyda hisoboti uchun Excel (3 ta varaq).

    summary: get_sales_summary() natijasi
    by_category: get_sales_by_category() natijasi
    details: get_sales_details() natijasi
    """
    wb = Workbook()

    # ===== Varaq 1: Umumiy =====
    ws = wb.active
    ws.title = "Umumiy"

    ws.merge_cells("A1:D1")
    ws["A1"] = f"💰 SOTUV HISOBOTI: {date_from} — {date_to}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    revenue = float(summary.get('revenue', 0) or 0)
    cost = float(summary.get('cost', 0) or 0)
    profit = float(summary.get('profit', 0) or 0)
    margin = float(summary.get('margin_percent', 0) or 0)

    rows = [
        ("Buyurtmalar soni", summary.get('orders_count', 0), None),
        ("Sotilgan dona", summary.get('units_sold', 0), None),
        ("Tushum (sotuv)", revenue, MONEY_FMT),
        ("Tannarx (xarid)", cost, MONEY_FMT),
        ("Sof foyda", profit, MONEY_FMT),
        ("Marja %", round(margin, 2), '0.00"%"'),
        ("To'liq tushgan pul", summary.get('paid_received', 0), MONEY_FMT),
        ("Qarzga ketgan", summary.get('debt_added', 0), MONEY_FMT),
    ]
    for i, (label, val, fmt) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = SUB_HEADER_FILL
        ws.cell(row=i, column=1).border = BORDER
        cell = ws.cell(row=i, column=2, value=val)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
        # Foyda qizil/yashil
        if label == "Sof foyda":
            cell.font = Font(bold=True, color="00B050" if profit >= 0 else "C00000", size=12)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 24

    # ===== Varaq 2: Kategoriya bo'yicha =====
    ws2 = wb.create_sheet("Kategoriya bo'yicha")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "📂 KATEGORIYALAR BO'YICHA SOTUV"
    ws2["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws2["A1"].fill = HEADER_FILL
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 24

    cat_headers = ["Kategoriya", "Buyurtmalar", "Dona", "Tushum", "Tannarx", "Foyda"]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    r = 4
    for c in by_category:
        cat_name = f"{c.get('category_icon', '📦')} {c.get('category_name', '—')}"
        cells = [
            cat_name,
            c.get('orders_count', 0),
            c.get('units', 0),
            float(c.get('revenue', 0) or 0),
            float(c.get('cost', 0) or 0),
            float(c.get('revenue', 0) or 0) - float(c.get('cost', 0) or 0),
        ]
        for ci, val in enumerate(cells, 1):
            cell = ws2.cell(row=r, column=ci, value=val)
            cell.border = BORDER
            if ci >= 4:
                cell.number_format = MONEY_FMT
        r += 1

    _autosize(ws2)

    # ===== Varaq 3: Buyurtmalar tafsiloti =====
    ws3 = wb.create_sheet("Buyurtmalar")
    ws3.merge_cells("A1:K1")
    ws3["A1"] = "📋 BUYURTMALAR RO'YXATI"
    ws3["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws3["A1"].fill = HEADER_FILL
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 24

    det_headers = ["#", "Sana", "Mijoz", "Telefon", "Mahsulot", "Dona",
                   "Tushum", "Tannarx", "Foyda", "To'lov", "Holat"]
    for col_idx, h in enumerate(det_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    payment_status_names = {"paid": "To'liq", "debt": "Qarz", "partial": "Qisman"}

    r = 4
    for o in details:
        product_full = " ".join(filter(None, [
            o.get('brand_name'), o.get('model_name'), o.get('product_name')
        ])) or o.get('product_name', '—')
        revenue = float(o.get('total_price', 0) or 0)
        cost_val = float(o.get('cost_at_sale', 0) or 0)
        profit_val = revenue - cost_val
        created = (o.get('created_at') or "")[:16]
        cells = [
            o.get('id'),
            created,
            o.get('customer', '—'),
            o.get('customer_phone', '—'),
            product_full,
            o.get('quantity', 0),
            revenue,
            cost_val,
            profit_val,
            payment_status_names.get(o.get('payment_status'), '?'),
            o.get('status', '—'),
        ]
        for ci, val in enumerate(cells, 1):
            cell = ws3.cell(row=r, column=ci, value=val)
            cell.border = BORDER
            if ci in (7, 8, 9):
                cell.number_format = MONEY_FMT
        r += 1

    # Yakuniy qator
    if details:
        ws3.cell(row=r, column=1, value="JAMI:").font = TOTAL_FONT
        ws3.cell(row=r, column=1).fill = TOTAL_FILL
        for col_idx in (7, 8, 9):
            ws3.cell(row=r, column=col_idx, value=f"=SUM({get_column_letter(col_idx)}4:{get_column_letter(col_idx)}{r-1})")
            ws3.cell(row=r, column=col_idx).number_format = MONEY_FMT
            ws3.cell(row=r, column=col_idx).font = TOTAL_FONT
            ws3.cell(row=r, column=col_idx).fill = TOTAL_FILL

    _autosize(ws3)

    path = _temp_path(f"sotuv_{date_from}_{date_to}")
    wb.save(path)
    return path
