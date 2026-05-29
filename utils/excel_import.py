"""
Excel orqali mahsulotlarni ommaviy import qilish:
- shablon (template) yaratish
- yuklangan faylni o'qib, qatorlarga ajratish
"""
import os
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Ustun sarlavhalarini moslashtirish (kichik harf, har xil nomlar qabul qilinadi)
HEADER_MAP = {
    "kategoriya": "category", "category": "category", "tur": "category",
    "brend": "brand", "brand": "brand",
    "model": "model",
    "nomi": "name", "nom": "name", "mahsulot": "name", "name": "name", "turi": "name",
    "tannarx": "cost", "cost": "cost",
    "narx": "price", "sotish": "price", "price": "price", "sotish narxi": "price",
    "miqdor": "quantity", "soni": "quantity", "qty": "quantity",
    "quantity": "quantity", "dona": "quantity",
    "min": "min", "minimum": "min", "min miqdor": "min",
    "tavsif": "description", "izoh": "description", "description": "description",
}

CATEGORIES = ["Ekran", "Orqa krishka", "Batareya", "Kamera shisha", "Pastki plata", "Aksessuar"]

HEADERS = ["Kategoriya", "Brend", "Model", "Nomi", "Tannarx", "Narx", "Miqdor", "Min", "Tavsif"]


def generate_template() -> str:
    """To'ldirish uchun namuna shablon .xlsx yaratadi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E78")
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(1, i, h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    examples = [
        ["Ekran", "iPhone", "13", "OLED Original", 595000, 850000, 10, 3, "Original ekran"],
        ["Ekran", "iPhone", "13", "OLED Copy", 350000, 520000, 20, 5, ""],
        ["Batareya", "Samsung", "S22", "Original 4000mAh", 90000, 150000, 15, 5, ""],
        ["Orqa krishka", "iPhone", "14", "Original krishka", 70000, 130000, 8, 3, ""],
        ["Aksessuar", "", "", "Universal kabel Type-C", 8000, 15000, 100, 10, "Brend shart emas"],
    ]
    for r, row in enumerate(examples, 2):
        for i, v in enumerate(row, 1):
            ws.cell(r, i, v)

    widths = [16, 12, 12, 28, 12, 12, 10, 8, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Yo'riqnoma varag'i
    ws2 = wb.create_sheet("Yo'riqnoma")
    notes = [
        "QO'LLANMA",
        "",
        "1) 'Mahsulotlar' varag'idagi namuna qatorlarni o'chiring.",
        "2) O'z mahsulotlaringizni yozing.",
        "3) Faylni saqlab, botga yuboring.",
        "",
        "USTUNLAR:",
        "• Kategoriya (MAJBURIY) — quyidagilardan biri:",
        *[f"      - {c}" for c in CATEGORIES],
        "• Brend — Ekran/Krishka/Batareya/Kamera/Plata uchun MAJBURIY (mas: iPhone).",
        "          Aksessuar uchun BO'SH qoldiriladi.",
        "• Model — yuqoridagilar uchun MAJBURIY (mas: 13, S22). Aksessuar uchun bo'sh.",
        "• Nomi (MAJBURIY) — mahsulot turi (mas: OLED Original, Original 4000mAh).",
        "• Tannarx — sotib olingan narx (raqam). Bo'sh bo'lsa 0.",
        "• Narx (MAJBURIY) — sotish narxi (raqam, 0 dan katta).",
        "• Miqdor — nechta dona (raqam). Bo'sh bo'lsa 0.",
        "• Min — shu sondan kam qolsa 'tugab qolgan' deb belgilanadi. Bo'sh bo'lsa 3.",
        "• Tavsif — ixtiyoriy.",
        "",
        "MUHIM: bir xil Kategoriya + Brend + Model + Nomi bo'lsa,",
        "mavjud mahsulot YANGILANADI (narx/miqdor ustiga yoziladi), takrorlanmaydi.",
        "Yangi brend/model bo'lsa — avtomatik yaratiladi.",
    ]
    for r, t in enumerate(notes, 1):
        cell = ws2.cell(r, 1, t)
        if r == 1:
            cell.font = Font(bold=True, size=14)
        elif t.endswith(":") or t.startswith("MUHIM"):
            cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 72

    path = os.path.join(tempfile.gettempdir(), "import_shablon.xlsx")
    wb.save(path)
    return path


def parse_import_file(path: str):
    """
    Yuklangan faylni o'qiydi.
    Qaytaradi: (rows, fatal_error)
      rows  — list[dict] (xom qiymatlar, har birida 'row' = Excel qator raqami)
      fatal — None yoki xato matni (jadval tuzilishi noto'g'ri bo'lsa)
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [], f"Faylni o'qib bo'lmadi: {e}"

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], "Fayl bo'sh."

    colmap = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = HEADER_MAP.get(str(h).strip().lower())
        if key and key not in colmap:
            colmap[key] = idx

    missing = [x for x in ("category", "name", "price") if x not in colmap]
    if missing:
        wb.close()
        nice = {"category": "Kategoriya", "name": "Nomi", "price": "Narx"}
        return [], ("Kerakli ustunlar topilmadi: "
                    + ", ".join(nice[m] for m in missing)
                    + ".\nShablondan foydalaning.")

    def cell(raw, k):
        i = colmap.get(k)
        if i is None or i >= len(raw):
            return None
        return raw[i]

    out = []
    rownum = 1
    for raw in rows_iter:
        rownum += 1
        if raw is None:
            continue
        # Butunlay bo'sh qatorni o'tkazib yuboramiz
        if not any(cell(raw, k) not in (None, "") for k in
                   ("category", "name", "price", "brand", "model")):
            continue
        out.append({
            "row": rownum,
            "category": cell(raw, "category"),
            "brand": cell(raw, "brand"),
            "model": cell(raw, "model"),
            "name": cell(raw, "name"),
            "cost": cell(raw, "cost"),
            "price": cell(raw, "price"),
            "quantity": cell(raw, "quantity"),
            "min": cell(raw, "min"),
            "description": cell(raw, "description"),
        })
        if len(out) >= 3000:
            break

    wb.close()
    return out, None
